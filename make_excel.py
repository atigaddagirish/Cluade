#!/usr/bin/env python3
"""Generate IS801_HAT_Calculator.xlsx — fully formula-linked IS 801:1975 calc.

Every result is a live Excel formula chained from the INPUT cells (yellow).
Mirrors the JS engine in index.html (thin-walled polyline + all checks).
Run `python3 make_excel.py`, then verify against validate.py targets.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "IS801"

# ── styles ──
TITLE = Font(bold=True, size=14, color="1E3A5F")
H1 = Font(bold=True, size=11, color="FFFFFF")
H1F = PatternFill("solid", fgColor="1E3A5F")
INP = PatternFill("solid", fgColor="FFF2CC")
OUT = PatternFill("solid", fgColor="E2EFDA")
LBL = Font(size=10, color="444444")
VAL = Font(size=10, bold=True)
REF = Font(size=9, italic=True, color="888888")
thin = Side(style="thin", color="CCCCCC")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

def head(r, text):
    ws.cell(r, 1, text).font = H1
    for c in range(1, 6):
        ws.cell(r, c).fill = H1F

def put(r, label, formula_or_val, unit="", ref="", inp=False):
    ws.cell(r, 1, label).font = LBL
    c = ws.cell(r, 2, formula_or_val)
    c.font = VAL
    c.fill = INP if inp else OUT
    c.border = BOX
    ws.cell(r, 3, unit).font = LBL
    ws.cell(r, 4, ref).font = REF
    return c

for col, w in (("A", 44), ("B", 16), ("C", 10), ("D", 46), ("E", 4),
               ("F", 8), ("G", 9), ("H", 9), ("I", 9), ("J", 9), ("K", 10),
               ("L", 10), ("M", 12), ("N", 12), ("O", 12), ("P", 12), ("Q", 12)):
    ws.column_dimensions[col].width = w

ws.cell(1, 1, "IS 801:1975 — Trapezoidal Lipped HAT Purlin Design Check (formula-linked)").font = TITLE
ws.cell(2, 1, "Yellow cells = inputs. Everything else recalculates automatically.").font = REF

# ── INPUTS ──
head(4, "1. INPUTS")
put(5,  "Top flange B1", 40, "mm", inp=True)                     # B5
put(6,  "Overall height H", 70, "mm", inp=True)                  # B6
put(7,  "Web angle θ (from horizontal)", 75, "deg", inp=True)    # B7
put(8,  "Foot width F", 22, "mm", inp=True)                      # B8
put(9,  "Lip length Lp", 10, "mm", inp=True)                     # B9
put(10, "Lip angle β (from horizontal)", 45, "deg", inp=True)    # B10
put(11, "Thickness t", 0.9, "mm", inp=True)                      # B11
put(12, "Yield Fy", 550, "MPa", inp=True)                        # B12
put(13, "Modulus E", 200000, "MPa", inp=True)                    # B13
put(14, "Span L", 3.1, "m", inp=True)                            # B14
put(15, "Support condition", "Simply supported", "", "dropdown", inp=True)  # B15
dv = DataValidation(type="list", formula1='"Simply supported,2-span continuous,3+ span continuous"', allow_blank=False)
ws.add_data_validation(dv); dv.add(ws["B15"])
put(16, "Unbraced length Lb", 3100, "mm", inp=True)              # B16
put(17, "Moment gradient Cb", 1.0, "", "Cb=1 with axial (Cl.6.7)", inp=True)  # B17
put(18, "Bearing length N", 50, "mm", inp=True)                  # B18
put(19, "Inner bend radius r", 2, "mm", inp=True)                # B19
put(20, "Deflection limit divisor (L/x)", 180, "", inp=True)     # B20
put(21, "w net uplift (working)", 0.981, "kN/m", "WL−DL, service", inp=True)  # B21
put(22, "w gravity (working)", 0.147, "kN/m", inp=True)          # B22
put(23, "w weak-axis", 0.0, "kN/m", inp=True)                    # B23
put(24, "Wind +33% allowable increase", 0, "", "REMOVED per project practice — allowables at 0.6Fy", inp=True)  # B24 (kept at 0)

# ── DERIVED GEOMETRY ──
head(26, "2. DERIVED GEOMETRY")
put(27, "half = B1/2", "=B5/2", "mm")                            # B27
put(28, "run = H/tan(θ)", "=B6/TAN(RADIANS(B7))", "mm")          # B28
put(29, "wb = half+run (web base x)", "=B27+B28", "mm")          # B29
put(30, "fe = wb+F (foot end x)", "=B29+B8", "mm")               # B30
put(31, "lip dx = Lp·cosβ", "=B9*COS(RADIANS(B10))", "mm")       # B31
put(32, "lip dy = Lp·sinβ", "=B9*SIN(RADIANS(B10))", "mm")       # B32
put(33, "lt = fe+lipdx (lip tip x)", "=B30+B31", "mm")           # B33
put(34, "web slant = √(run²+H²)", "=SQRT(B28^2+B6^2)", "mm")     # B34

# ── SEGMENT TABLE (thin-walled polyline, 7 segments) ──
head(36, "3. SECTION PROPERTY ENGINE — thin-walled polyline (7 segments)")
hdr = ["seg", "x1", "y1", "x2", "y2", "L", "A=L·t", "xm", "ym", "φ (rad)",
       "A·ym", "Ix,own", "Iy,own", "A·(ym-ȳ)²", "Ixx contrib", "A·xm²", "Iyy contrib(x=0)"]
for j, h in enumerate(hdr):
    c = ws.cell(37, 1 + j, h); c.font = Font(bold=True, size=9); c.border = BOX
# vertex coordinate expressions (lipL→footL→webL→top→webR→footR→lipR)
VX = ["-$B$33", "-$B$30", "-$B$29", "-$B$27", "$B$27", "$B$29", "$B$30", "$B$33"]
VY = ["$B$32", "0", "0", "$B$6", "$B$6", "0", "0", "$B$32"]
SEG0 = 38  # first segment row
for i in range(7):
    r = SEG0 + i
    ws.cell(r, 1, i).border = BOX
    ws.cell(r, 2, f"={VX[i]}")
    ws.cell(r, 3, f"={VY[i]}")
    ws.cell(r, 4, f"={VX[i+1]}")
    ws.cell(r, 5, f"={VY[i+1]}")
    ws.cell(r, 6, f"=SQRT((D{r}-B{r})^2+(E{r}-C{r})^2)")                    # L
    ws.cell(r, 7, f"=F{r}*$B$11")                                            # A
    ws.cell(r, 8, f"=(B{r}+D{r})/2")                                         # xm
    ws.cell(r, 9, f"=(C{r}+E{r})/2")                                         # ym
    ws.cell(r, 10, f"=ATAN2(D{r}-B{r},E{r}-C{r})")                           # phi
    ws.cell(r, 11, f"=G{r}*I{r}")                                            # A·ym
    ws.cell(r, 12, f"=($B$11*F{r}^3/12)*SIN(J{r})^2+(F{r}*$B$11^3/12)*COS(J{r})^2")  # Ix own
    ws.cell(r, 13, f"=($B$11*F{r}^3/12)*COS(J{r})^2+(F{r}*$B$11^3/12)*SIN(J{r})^2")  # Iy own
    ws.cell(r, 14, f"=G{r}*(I{r}-$B$48)^2")                                  # A(ym-ybar)^2
    ws.cell(r, 15, f"=L{r}+N{r}")                                            # Ixx contrib
    ws.cell(r, 16, f"=G{r}*H{r}^2")                                          # A·xm² (xc=0 symmetric)
    ws.cell(r, 17, f"=M{r}+P{r}")                                            # Iyy contrib about x=0
SEGE = SEG0 + 6  # 44

head(46, "4. SECTION PROPERTIES (gross)")
put(47, "Area A", f"=SUM(G{SEG0}:G{SEGE})", "mm²", "target 224.0")            # B47
put(48, "ȳ (from foot line)", f"=SUM(K{SEG0}:K{SEGE})/B47", "mm", "target 31.91")  # B48
put(49, "Ixx", f"=SUM(O{SEG0}:O{SEGE})", "mm⁴", "target 161637")              # B49
put(50, "Iyy", f"=SUM(Q{SEG0}:Q{SEGE})", "mm⁴", "target 295330 (xc=0 by symmetry)")  # B50
put(51, "y_top = H − ȳ", "=B6-B48", "mm")                                     # B51
put(52, "y_bot = ȳ (foot line lowest)", "=B48", "mm")                         # B52
put(53, "Ztop = Ixx/y_top", "=B49/B51", "mm³", "target 4244")                 # B53
put(54, "Zbot = Ixx/y_bot", "=B49/B52", "mm³", "target 5065")                 # B54
put(55, "Zweak = Iyy/lt", "=B50/B33", "mm³")                                  # B55
put(56, "J = Σ(L·t³)/3", f"=SUM(F{SEG0}:F{SEGE})*B11^3/3", "mm⁴")             # B56
put(57, "Acf (feet+lips: segs 0,1,5,6)", f"=G{SEG0}+G{SEG0+1}+G{SEG0+5}+G{SEG0+6}", "mm²")  # B57
put(58, "Iycf about x=0", f"=Q{SEG0}+Q{SEG0+1}+Q{SEG0+5}+Q{SEG0+6}", "mm⁴")   # B58
put(59, "r_ycf = √(Iycf/Acf)", "=SQRT(B58/B57)", "mm", "target 55.0")         # B59

# ── LOADS & BENDING ──
head(61, "5. MOMENTS & BENDING — Cl.6.1.1")
put(62, "M uplift = w·L²/8", "=B21*B14^2/8", "kN·m", "wL²/8: exact 2-span support M, conservative 3+")  # B62
put(63, "M gravity = w·L²/8", "=B22*B14^2/8", "kN·m")                          # B63
put(64, "M weak = w·L²/8", "=B23*B14^2/8", "kN·m")                             # B64
put(65, "Fb = 0.6·Fy", "=0.6*B12", "MPa", "Cl.6.1.1")                          # B65
put(66, "Fb design (no wind increase)", "=B65", "MPa")                            # B66
put(67, "f_top (tension, uplift) = M/Ztop", "=B62*10^6/B53", "MPa")            # B67
put(68, "f_bot (compression, uplift) = M/Zbot", "=B62*10^6/B54", "MPa")        # B68
put(69, "f_strong = max(f_top, f_bot)", "=MAX(B67,B68)", "MPa")                # B69
put(70, "f_weak = M_weak/Zweak", "=B64*10^6/B55", "MPa")                       # B70
put(71, "UR bending (biaxial) = (f_strong+f_weak)/Fb", "=(B69+B70)/B66", "", "Cl.6.1.1")  # B71
put(72, "f_top gravity = M_grav/Ztop", "=B63*10^6/B53", "MPa")                 # B72
put(73, "UR gravity bending = f/Fb", "=B72/B65", "", "no +⅓ (DL+LL)")          # B73

# ── EFFECTIVE WIDTH ──
head(75, "6. EFFECTIVE WIDTH — Cl.5.2.1.1 (AISI unified Winter λ-form)")
put(76, "f for eff. width = max(f_bot,1)", "=MAX(B68,1)", "MPa")               # B76
put(77, "λ foot (k=4) =(1.052/√4)(F/t)√(f/E)", "=(1.052/SQRT(4))*(B8/B11)*SQRT(B76/B13)", "", "≤0.673 fully eff.")  # B77
put(78, "λ lip (k=0.43)", "=(1.052/SQRT(0.43))*(B9/B11)*SQRT(B76/B13)", "")    # B78
put(79, "λ web (k=24, slant)", "=(1.052/SQRT(24))*(B34/B11)*SQRT(B76/B13)", "")  # B79
put(80, "ρ foot", "=IF(B77<=0.673,1,(1-0.22/B77)/B77)", "")                    # B80
put(81, "ρ lip", "=IF(B78<=0.673,1,(1-0.22/B78)/B78)", "")                     # B81
put(82, "ρ web", "=IF(B79<=0.673,1,(1-0.22/B79)/B79)", "")                     # B82
put(83, "UR eff.width = max(λ)/0.673", "=MAX(B77:B79)/0.673", "")              # B83

head(85, "7. EDGE STIFFENER — Cl.5.2.2.1")
put(86, "Is = Lp³·t·sin²β/12", "=B9^3*B11*SIN(RADIANS(B10))^2/12", "mm⁴")      # B86
put(87, "Fy (ksi)", "=B12/6.895", "ksi")                                       # B87
put(88, "Imin = max(1.83t⁴√((F/t)²−4000/Fy_ksi), 9.2t⁴)",
        "=MAX(1.83*B11^4*SQRT(MAX((B8/B11)^2-4000/B87,0)),9.2*B11^4)", "mm⁴")  # B88
put(89, "UR stiffener = Imin/Is", "=B88/B86", "", "<1 → adequate")             # B89

# ── LTB ──
head(91, "8. LTB — Cl.6.3 (two methods, governing = min)")
put(92, "(a) Fe = Cb·π²E/(Lb/rycf)²", "=B17*PI()^2*B13/(B16/B59)^2", "MPa")    # B92
put(93, "(a) Fb: ≥2.78Fy→0.6Fy; ≥0.56Fy→(10Fy/9)(1−Fy/5.04Fe); else Fe/1.67",
        "=MIN(0.6*B12,IF(B92>=2.78*B12,0.6*B12,IF(B92>=0.56*B12,(10*B12/9)*(1-B12/(5.04*B92)),B92/1.67)))",
        "MPa", "compression-flange method")                                     # B93
put(94, "(b) Sxc = Zbot (Ix/dist to comp. fibre)", "=B54", "mm³", "literal Cl.6.3")  # B94
# Iyc = compression PORTION below NA: feet+lips (B58) + 2 × web sub-segment (y=0→ȳ)
# web sub-segment: Lsub=(ȳ/H)·slant; x from wb to wb−run·(ȳ/H); about x=0
put(95, "(b) Iyc = feet+lips + 2×web below NA",
        "=B58+2*(($B$11*((B48/B6)*B34)^3/12)*(B28/B34)^2"
        "+(((B48/B6)*B34)*$B$11^3/12)*(B6/B34)^2"
        "+((B48/B6)*B34)*$B$11*((B29+(B29-B28*(B48/B6)))/2)^2)",
        "mm⁴", "comp. portion about gravity axis ∥ web")                        # B95
put(96, "(b) X = Lb²·Sxc/(H·Iyc)", "=B16^2*B94/(B6*B95)", "")                  # B96
put(97, "(b) B = 0.36π²ECb/Fy", "=0.36*PI()^2*B13*B17/B12", "")                # B97
put(98, "(b) C = 1.8π²ECb/Fy", "=1.8*PI()^2*B13*B17/B12", "")                  # B98
put(99, "(b) Fb: X≤B→0.6Fy; B<X<C→⅔Fy−Fy²X/(5.4π²ECb); X≥C→0.6π²ECb/X",
        "=MIN(0.6*B12,IF(B96<=B97,0.6*B12,IF(B96<B98,(2/3)*B12-(B12^2/(5.4*PI()^2*B13*B17))*B96,0.6*PI()^2*B13*B17/B96)))",
        "MPa")                                                                  # B99
put(100, "Fb_LTB = min(a,b)", "=MIN(B93,B99)", "MPa")                          # B100
put(101, "Fb_LTB design (no wind increase)", "=B100", "MPa")                      # B101
put(102, "UR LTB = f_bot/Fb_LTB", "=B68/B101", "")                                # B102

# ── SHEAR ──
head(104, "9. SHEAR — Cl.6.4.1")
put(105, "h/t (web slant)", "=B34/B11", "")                                    # B105
put(106, "h/t limit = 4590/√Fy (kgf/cm²)", "=4590/SQRT(B115)", "")             # B106
put(107, "Fv (kgf/cm²): ≤lim→min(0.4Fy, 1275√Fy/(h/t)); else 5.85e6/(h/t)²",
         "=IF(B105<=B106,MIN(0.4*B115,1275*SQRT(B115)/B105),5850000/B105^2)", "kgf/cm²", "literal Cl.6.4.1")  # B107
put(108, "Fv (MPa)", "=B107*0.0980665", "MPa")                                  # B108
put(109, "Vcap = 2·Fv·H·t", "=2*B108*B6*B11/1000", "kN")                       # B109
put(110, "V = max(w)·L/2", "=MAX(ABS(B21),B22)*B14/2", "kN")                   # B110
put(111, "UR shear = V/Vcap", "=B110/B109", "")                                # B111

# ── WEB CRIPPLING ──
head(113, "10. WEB CRIPPLING — Cl.6.5 (literal MKS: t in cm, Fy in kgf/cm²)")
put(114, "t (cm)", "=B11/10", "cm")                                            # B114
put(115, "Fy (kgf/cm²)", "=B12/0.0980665", "kgf/cm²")                          # B115
put(116, "k = Fy/2320", "=B115/2320", "")                                      # B116
put(117, "stress end k(1.33−0.33k); int k(1.22−0.22k)", "=B116*(1.33-0.33*B116)", "", "end term")  # B117
ws.cell(117, 5, "=B116*(1.22-0.22*B116)").font = VAL
ws.cell(117, 5).fill = OUT; ws.cell(117, 5).border = BOX
ws.cell(117, 6, "int term (E117)").font = REF
put(118, "r-factor end = if r≤t:1 else 1.15−0.15r/t", "=IF(B19<=B11,1,1.15-0.15*B19/B11)", "")  # B118
put(119, "r-factor int = if r≤t:1 else 1.06−0.06r/t", "=IF(B19<=B11,1,1.06-0.06*B19/B11)", "")  # B119
put(120, "P_end = 70t²[98+4.2(N/t)−0.022(N/t)(h/t)−0.011(h/t)]·st·rf ×2 webs",
         "=MAX(0,70*B114^2*(98+4.2*(B18/B11)-0.022*(B18/B11)*(B105)-0.011*B105)*B117*B118)*2*9.80665/1000",
         "kN")                                                                  # B120
put(121, "P_int = 70t²[305+2.3(N/t)−0.009(N/t)(h/t)−0.5(h/t)]·st·rf ×2 webs",
         "=MAX(0,70*B114^2*(305+2.3*(B18/B11)-0.009*(B18/B11)*B105-0.5*B105)*E117*B119)*2*9.80665/1000",
         "kN")                                                                  # B121
put(122, "R_end (gravity) coeff·w·L", '=IF(B15="2-span continuous",0.375,IF(B15="3+ span continuous",0.4,0.5))*B22*B14', "kN")  # B122
put(123, "R_int (gravity)", '=IF(B15="2-span continuous",1.25,IF(B15="3+ span continuous",1.1,0))*B22*B14', "kN")  # B123
put(124, "UR web crippling = max(Rend/Pend, Rint/Pint)",
         "=MAX(B122/B120,IF(B123>0,B123/B121,0))", "", "uplift reaction → bolt check, not web")  # B124

# ── DEFLECTION ──
head(126, "11. DEFLECTION — L/" + "x limit")
put(127, "k coeff: ss 5/384; 2-span 0.00542; 3+ 0.0069",
         '=IF(B15="2-span continuous",0.00542,IF(B15="3+ span continuous",0.0069,5/384))', "")  # B127
put(128, "δ strong = k·w·(1000L)⁴/(E·Ixx)", "=B127*ABS(B21)*(B14*1000)^4/(B13*B49)", "mm", "1 kN/m = 1 N/mm")  # B128
put(129, "δ weak = k·w_weak·(1000L)⁴/(E·Iyy)", "=B127*ABS(B23)*(B14*1000)^4/(B13*B50)", "mm")  # B129
put(130, "δ resultant = √(δs²+δw²)", "=SQRT(B128^2+B129^2)", "mm")             # B130
put(131, "δ allowable = 1000L/divisor", "=B14*1000/B20", "mm")                 # B131
put(132, "UR deflection = δ/δ_allow", "=B130/B131", "")                        # B132

# ── ADDITIONAL WEB & RATIO CHECKS ──
head(134, "12. WEB BENDING Cl.6.4.2 / COMBINED Cl.6.4.3 / FLAT-WIDTH Cl.5.2.4")
put(135, "Fbw = min(36.56e6/(h/t)² kgf/cm², 0.6Fy)", "=MIN(36560000/B105^2*0.0980665,0.6*B12)", "MPa", "Cl.6.4.2")  # B135
put(136, "Fbw design (no wind increase)", "=B135", "MPa")                          # B136
put(137, "UR web bending = f_strong/Fbw", "=B69/B136", "")                      # B137
put(138, "fv actual = V/(2·H·t)", "=B110*1000/(2*B6*B11)", "MPa")               # B138
put(139, "UR combined = √((f/Fbw)²+(fv/Fv)²)", "=SQRT((B69/B136)^2+(B138/B108)^2)", "", "Cl.6.4.3")  # B139
put(140, "UR flange w/t ≤ 60", "=(B8/B11)/60", "", "Cl.5.2.4")                  # B140
put(141, "UR web h/t ≤ 150", "=B105/150", "", "Cl.5.2.4")                       # B141

# ── SUMMARY ──
head(143, "13. SUMMARY OF CHECKS")
checks = [
    ("Effective width (max λ/0.673)", "B83", "Cl.5.2.1.1"),
    ("Edge stiffener Imin/Is", "B89", "Cl.5.2.2.1"),
    ("Bending biaxial (uplift)", "B71", "Cl.6.1.1"),
    ("Bending gravity", "B73", "Cl.6.1.1"),
    ("LTB (governing of 2 methods)", "B102", "Cl.6.3"),
    ("Shear", "B111", "Cl.6.4.1"),
    ("Web crippling (gravity bearing)", "B124", "Cl.6.5"),
    ("Web bending", "B137", "Cl.6.4.2"),
    ("Combined bending + shear", "B139", "Cl.6.4.3"),
    ("Flange w/t ≤ 60", "B140", "Cl.5.2.4"),
    ("Web h/t ≤ 150", "B141", "Cl.5.2.4"),
    ("Deflection", "B132", "serviceability"),
]
ws.cell(144, 1, "Check").font = Font(bold=True)
ws.cell(144, 2, "UR").font = Font(bold=True)
ws.cell(144, 3, "Status").font = Font(bold=True)
ws.cell(144, 4, "Clause").font = Font(bold=True)
for i, (name, cell, cl) in enumerate(checks):
    r = 145 + i
    ws.cell(r, 1, name).font = LBL
    c = ws.cell(r, 2, f"={cell}"); c.number_format = "0.0%"; c.font = VAL; c.border = BOX
    s = ws.cell(r, 3, f'=IF({cell}>=1,"FAIL",IF({cell}>=0.9,"MARGINAL","PASS"))')
    s.font = VAL; s.border = BOX
    ws.cell(r, 4, cl).font = REF
r = 145 + len(checks)
ws.cell(r, 1, "OVERALL (max UR)").font = Font(bold=True, size=11)
c = ws.cell(r, 2, "=MAX(" + ",".join(x[1] for x in checks) + ")")
c.number_format = "0.0%"; c.font = Font(bold=True, size=11); c.border = BOX
s = ws.cell(r, 3, f'=IF(B{r}>=1,"FAIL",IF(B{r}>=0.9,"MARGINAL","PASS"))')
s.font = Font(bold=True, size=11); s.border = BOX

wb.save("IS801_HAT_Calculator.xlsx")
print("written IS801_HAT_Calculator.xlsx")
